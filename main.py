# --- Imports and Setup ---
import os
import json
import io
from dotenv import load_dotenv
from typing import Dict
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

# Load environment variables from .env file
load_dotenv()

# Get the OpenAI API key from the environment
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# Check if the API key is set
if not OPENAI_API_KEY:
    raise HTTPException(status_code=500, detail="OpenAI API key not found.")

# Initialize the OpenAI client
client = OpenAI(api_key=OPENAI_API_KEY)

# Initialize FastAPI app
app = FastAPI()

# Add CORS middleware to allow requests from your frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# --- Knowledge Base and Utility Functions ---

# Define the knowledge base with financial assumptions
knowledge_base = {
    "large_customer": {
        "revenue_per_customer": 16500,
        "sales_per_person_per_month": 1.5
    },
    "smb_customer": {
        "marketing_spend": 200000,
        "cac": 1500,
        "conversion_rate": 0.45,
        "revenue_per_customer": 5000
    },
    "sales_team": {
        "initial_salespeople": 2,
        "monthly_hires": 1
    }
}

# The function to generate the Excel file
def generate_forecast(params: dict) -> io.BytesIO:
    months = params.get("months", 6)
    start = params.get("start", "Jan 2025")

    kb = knowledge_base.copy()
    if "marketing_spend" in params:
        kb["smb_customer"]["marketing_spend"] = params["marketing_spend"]
    if "cac" in params:
        kb["smb_customer"]["cac"] = params["cac"]
    if "conversion_rate" in params:
        kb["smb_customer"]["conversion_rate"] = params["conversion_rate"]
    if "revenue_per_customer" in params:
        kb["smb_customer"]["revenue_per_customer"] = params["revenue_per_customer"]
    if "initial_salespeople" in params:
        kb["sales_team"]["initial_salespeople"] = params["initial_salespeople"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forecast"

    headers = [""] + [f"M{i+1}" for i in range(months)]
    ws.append(headers)

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for col in range(1, months + 2):
        ws.cell(1, col).font = bold_font
        ws.cell(1, col).alignment = center_align

    salespeople = kb["sales_team"]["initial_salespeople"]
    sales_per_person = kb["large_customer"]["sales_per_person_per_month"]
    large_rev_per_customer = kb["large_customer"]["revenue_per_customer"]

    marketing = kb["smb_customer"]["marketing_spend"]
    cac = kb["smb_customer"]["cac"]
    conv = kb["smb_customer"]["conversion_rate"]
    smb_rev_per_customer = kb["smb_customer"]["revenue_per_customer"]

    salespeople_list, large_customers_list, smb_customers_list = [], [], []
    large_revenue_list, smb_revenue_list, total_revenue_list = [], [], []
    sales_enquiries_list, paying_smb_list = [], []

    cumulative_large = 0
    cumulative_smb = 0

    for m in range(months):
        new_large = salespeople * sales_per_person
        cumulative_large += new_large
        leads = marketing / cac
        new_smb = leads * conv
        cumulative_smb += new_smb
        large_revenue = cumulative_large * large_rev_per_customer
        smb_revenue = cumulative_smb * smb_rev_per_customer
        total_revenue = large_revenue + smb_revenue
        enquiries = marketing / cac
        paying_smb = new_smb
        salespeople_list.append(salespeople)
        large_customers_list.append(cumulative_large)
        smb_customers_list.append(cumulative_smb)
        large_revenue_list.append(large_revenue)
        smb_revenue_list.append(smb_revenue)
        total_revenue_list.append(total_revenue)
        sales_enquiries_list.append(enquiries)
        paying_smb_list.append(new_smb)

        salespeople += kb["sales_team"]["monthly_hires"]

    rows = [
        ("# of sales people", salespeople_list, " #"),
        ("# of large customer accounts they can sign per month/ sales person",
         [sales_per_person] * months, " #"),
        ("# of large customer accounts onboarded per month",
         [round(salespeople_list[i] * sales_per_person) for i in range(months)], " #"),
        ("Cumulative # of paying customers",
         [round(x) for x in large_customers_list], " #"),
        ("Average revenue per customer",
         [large_rev_per_customer] * months, "₹ per month"),
        ("Digital Marketing spend per month",
         [marketing] * months, "₹ per month"),
        ("Average CAC",
         [cac] * months, "₹ per customer"),
        ("# of sales enquiries",
         [round(x) for x in sales_enquiries_list], " #"),
        ("% conversions from demo to sign ups",
         [conv] * months, "%"),
        ("# of paying customers onboarded",
         [round(x) for x in paying_smb_list], " #"),
        ("Cumulative number of paying customers",
         [round(x) for x in smb_customers_list], " #"),
        ("Average revenue per customer",
         [smb_rev_per_customer] * months, "₹ per customer"),
        ("Revenue from large clients",
         [round(x) for x in large_revenue_list], "₹"),
        ("Revenue from small and medium clients",
         [round(x) for x in smb_revenue_list], "₹"),
        ("Total Revenues",
         [round(x) for x in total_revenue_list], "₹"),
    ]

    for label, values, unit in rows:
        row = [label] + values
        ws.append(row)

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=months+1):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 60
    for col in range(2, months+2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- API Endpoints ---

class PromptInput(BaseModel):
    prompt: str

@app.post("/forecast_from_prompt")
async def forecast_from_prompt(data: PromptInput):
    try:
        system_msg = {
            "role": "system",
            "content": (
                "You are a helpful finance assistant. "
                "Extract structured JSON forecast parameters from the user prompt. "
                "Keys allowed: months (int), start (str), marketing_spend (int), "
                "cac (int), conversion_rate (float), revenue_per_customer (int), "
                "initial_salespeople (int). "
                "If not provided, leave them out."
            )
        }
        user_msg = {"role": "user", "content": data.prompt}

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[system_msg, user_msg],
            response_format={"type": "json_object"}
        )

        params = json.loads(response.choices[0].message.content)
        excel_file = generate_forecast(params)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=forecast.xlsx"}
        )

    except Exception as e:
        # Return a JSON error response that the frontend can handle
        raise HTTPException(status_code=500, detail=str(e))
